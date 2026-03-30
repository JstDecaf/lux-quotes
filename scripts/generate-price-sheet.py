#!/usr/bin/env python3
"""
Generate a Reseller Price Sheet (XLSX) for a LUX quote.
Usage: python3 generate-price-sheet.py <quote_id> [db_path] [output_path]
"""

import sys
import os
import sqlite3
import tempfile
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


def calculate_line_item(item, settings):
    """Replicate calculations.ts logic exactly."""
    if item["is_free"]:
        return {
            "usd_subtotal": 0,
            "aud_cost": 0,
            "aud_sell_ex_gst": 0,
            "gst": 0,
            "aud_sell_inc_gst": 0,
            "gross_profit": 0,
            "reseller_sell_ex_gst": 0,
            "reseller_gst": 0,
            "reseller_sell_inc_gst": 0,
            "reseller_profit": 0,
        }

    margin = item["margin_override"] if item["margin_override"] is not None else settings["default_margin"]
    reseller_margin = item["reseller_margin_override"] if item["reseller_margin_override"] is not None else settings["default_reseller_margin"]

    is_local = bool(item["is_local"])
    usd_subtotal = 0 if is_local else item["qty"] * (item["usd_unit_price"] or 0)
    aud_cost = (item["aud_local_cost"] or 0) if is_local else (usd_subtotal / settings["fx_rate"] if settings["fx_rate"] > 0 else 0)

    # LUX sell price (what LUX charges the reseller)
    aud_sell_ex_gst = aud_cost / (1 - margin) if margin < 1 else aud_cost
    gst = aud_sell_ex_gst * settings["gst_rate"]
    aud_sell_inc_gst = aud_sell_ex_gst + gst
    gross_profit = aud_sell_ex_gst - aud_cost

    # Reseller sell price (what reseller charges end client)
    reseller_sell_ex_gst = aud_sell_ex_gst / (1 - reseller_margin) if reseller_margin < 1 else aud_sell_ex_gst
    reseller_gst = reseller_sell_ex_gst * settings["gst_rate"]
    reseller_sell_inc_gst = reseller_sell_ex_gst + reseller_gst
    reseller_profit = reseller_sell_ex_gst - aud_sell_ex_gst

    return {
        "usd_subtotal": usd_subtotal,
        "aud_cost": aud_cost,
        "aud_sell_ex_gst": aud_sell_ex_gst,
        "gst": gst,
        "aud_sell_inc_gst": aud_sell_inc_gst,
        "gross_profit": gross_profit,
        "reseller_sell_ex_gst": reseller_sell_ex_gst,
        "reseller_gst": reseller_gst,
        "reseller_sell_inc_gst": reseller_sell_inc_gst,
        "reseller_profit": reseller_profit,
    }


def main():
    quote_id = int(sys.argv[1])
    db_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.path.dirname(__file__), "..", "data", "lux-quotes.db")
    output_path = sys.argv[3] if len(sys.argv) > 3 else None

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    # Fetch quote with project/client info
    quote = conn.execute("""
        SELECT q.*, p.name as project_name, c.name as client_name, c.contact_name, c.contact_email
        FROM quotes q
        LEFT JOIN projects p ON q.project_id = p.id
        LEFT JOIN clients c ON p.client_id = c.id
        WHERE q.id = ?
    """, (quote_id,)).fetchone()

    if not quote:
        print(f"ERROR: Quote {quote_id} not found", file=sys.stderr)
        sys.exit(1)

    items = conn.execute("""
        SELECT * FROM quote_line_items
        WHERE quote_id = ?
        ORDER BY sort_order
    """, (quote_id,)).fetchall()

    settings = {
        "fx_rate": quote["fx_rate"],
        "default_margin": quote["default_margin"],
        "default_reseller_margin": quote["default_reseller_margin"],
        "gst_rate": quote["gst_rate"],
        "deposit_pct": quote["deposit_pct"],
        "second_tranche_pct": quote["second_tranche_pct"],
    }

    # Calculate all line items
    calculated = []
    for item in items:
        calc = calculate_line_item(dict(item), settings)
        calculated.append(calc)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price Sheet"

    # Styles
    navy = "0D1B2A"
    red = "DB412B"
    white_font = Font(name="Inter", color="FFFFFF", bold=True, size=10)
    header_fill = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
    red_fill = PatternFill(start_color=red, end_color=red, fill_type="solid")
    title_font = Font(name="Archivo", bold=True, size=20, color=navy)
    subtitle_font = Font(name="Archivo", bold=True, size=12, color=navy)
    label_font = Font(name="Inter", size=10, color="4A5568")
    value_font = Font(name="Inter", size=10, color="2D3748")
    bold_value_font = Font(name="Inter", size=10, color="2D3748", bold=True)
    currency_fmt = '#,##0.00'
    pct_fmt = '0.0%'
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    totals_border = Border(
        top=Side(style="medium", color=navy),
        bottom=Side(style="double", color=navy),
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
    )

    # Column widths
    col_widths = {
        "A": 28,  # Item Name
        "B": 32,  # Description
        "C": 8,   # Unit
        "D": 8,   # Qty
        "E": 14,  # Ex-GST (per unit)
        "F": 14,  # Ex-GST (total)
        "G": 12,  # GST
        "H": 14,  # Inc-GST (total)
        "I": 12,  # Suggested Margin %
        "J": 16,  # Suggested Reseller Ex-GST
        "K": 16,  # Suggested Reseller Inc-GST
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    row = 1

    # === HEADER SECTION ===
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "LUX LED Solutions"
    ws[f"A{row}"].font = Font(name="Archivo", bold=True, size=16, color=red)
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "RESELLER PRICE SHEET"
    ws[f"A{row}"].font = title_font
    ws.row_dimensions[row].height = 30
    row += 1

    row += 1  # blank row

    # Quote ref and date
    ws[f"A{row}"] = "Quote Ref:"
    ws[f"A{row}"].font = label_font
    ws[f"B{row}"] = quote["quote_number"]
    ws[f"B{row}"].font = bold_value_font

    ws[f"D{row}"] = "Date:"
    ws[f"D{row}"].font = label_font
    ws[f"E{row}"] = datetime.now().strftime("%d %b %Y")
    ws[f"E{row}"].font = bold_value_font
    row += 1

    ws[f"A{row}"] = "Quote Name:"
    ws[f"A{row}"].font = label_font
    ws[f"B{row}"] = quote["name"]
    ws[f"B{row}"].font = value_font
    row += 1

    row += 1  # blank

    # Client/project info
    ws[f"A{row}"] = "Client:"
    ws[f"A{row}"].font = label_font
    ws[f"B{row}"] = quote["client_name"] or "N/A"
    ws[f"B{row}"].font = bold_value_font

    ws[f"D{row}"] = "Project:"
    ws[f"D{row}"].font = label_font
    ws[f"E{row}"] = quote["project_name"] or "N/A"
    ws[f"E{row}"].font = bold_value_font
    row += 1

    if quote["contact_name"]:
        ws[f"A{row}"] = "Contact:"
        ws[f"A{row}"].font = label_font
        ws[f"B{row}"] = quote["contact_name"]
        ws[f"B{row}"].font = value_font
        row += 1

    row += 1  # blank

    # === TABLE HEADER ===
    headers = [
        "Item Name",
        "Description",
        "Unit",
        "Qty",
        "Ex-GST\n(per unit)",
        "Ex-GST\n(total)",
        "GST",
        "Inc-GST\n(total)",
        "Suggested\nReseller Margin",
        "Suggested Reseller\nEx-GST",
        "Suggested Reseller\nInc-GST",
    ]

    table_start_row = row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[row].height = 36
    row += 1

    # === TABLE DATA ===
    for i, item in enumerate(items):
        calc = calculated[i]
        qty = item["qty"] or 0

        # Per-unit LUX sell ex-GST
        per_unit_ex_gst = calc["aud_sell_ex_gst"] / qty if qty > 0 else 0

        # Reseller margin used
        reseller_margin = item["reseller_margin_override"] if item["reseller_margin_override"] is not None else settings["default_reseller_margin"]

        row_data = [
            item["item_name"],
            item["description"] or "",
            item["unit"],
            qty,
            per_unit_ex_gst,
            calc["aud_sell_ex_gst"],
            calc["gst"],
            calc["aud_sell_inc_gst"],
            reseller_margin,
            calc["reseller_sell_ex_gst"],
            calc["reseller_sell_inc_gst"],
        ]

        # Alternate row shading
        row_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid") if i % 2 == 0 else None

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = value_font
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill

            # Formatting
            if col_idx in (1, 2):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx == 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = '0'
            elif col_idx in (5, 6, 7, 8, 10, 11):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = currency_fmt
            elif col_idx == 9:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = pct_fmt

        row += 1

    # === TOTALS ROW ===
    totals_data = {
        1: "TOTALS",
        5: sum(c["aud_sell_ex_gst"] / (items[i]["qty"] or 1) for i, c in enumerate(calculated)),
        6: sum(c["aud_sell_ex_gst"] for c in calculated),
        7: sum(c["gst"] for c in calculated),
        8: sum(c["aud_sell_inc_gst"] for c in calculated),
        10: sum(c["reseller_sell_ex_gst"] for c in calculated),
        11: sum(c["reseller_sell_inc_gst"] for c in calculated),
    }

    for col_idx in range(1, 12):
        cell = ws.cell(row=row, column=col_idx)
        cell.border = totals_border
        cell.font = Font(name="Inter", bold=True, size=10, color=navy)

        if col_idx in totals_data:
            cell.value = totals_data[col_idx]
        if col_idx in (5, 6, 7, 8, 10, 11):
            cell.number_format = currency_fmt
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif col_idx == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[row].height = 24
    row += 2

    # === PAYMENT SCHEDULE ===
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "Payment Schedule"
    ws[f"A{row}"].font = subtitle_font
    row += 1

    deposit_pct = settings["deposit_pct"]
    second_pct = settings["second_tranche_pct"]
    balance_pct = 1 - deposit_pct - second_pct

    total_reseller_inc_gst = sum(c["reseller_sell_inc_gst"] for c in calculated)

    schedule = [
        (f"Deposit ({deposit_pct * 100:.0f}%)", total_reseller_inc_gst * deposit_pct),
        (f"2nd Payment ({second_pct * 100:.0f}%)", total_reseller_inc_gst * second_pct),
        (f"Balance ({balance_pct * 100:.0f}%)", total_reseller_inc_gst * balance_pct),
    ]

    for label, amount in schedule:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = label_font
        ws[f"B{row}"] = amount
        ws[f"B{row}"].font = bold_value_font
        ws[f"B{row}"].number_format = '$#,##0.00'
        row += 1

    row += 1
    ws[f"A{row}"] = f"Total (Inc GST)"
    ws[f"A{row}"].font = Font(name="Inter", bold=True, size=10, color=navy)
    ws[f"B{row}"] = total_reseller_inc_gst
    ws[f"B{row}"].font = Font(name="Inter", bold=True, size=12, color=red)
    ws[f"B{row}"].number_format = '$#,##0.00'

    row += 2

    # === FOOTER ===
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "LUX LED Solutions | Prices in AUD | Valid 30 days"
    ws[f"A{row}"].font = Font(name="Inter", size=8, color="9CA3AF", italic=True)
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

    # Set print area
    ws.print_area = f"A1:K{row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # Save
    if output_path is None:
        output_path = os.path.join(tempfile.gettempdir(), f"LUX-PriceSheet-{quote['quote_number']}.xlsx")

    wb.save(output_path)
    print(output_path)
    conn.close()


if __name__ == "__main__":
    main()
